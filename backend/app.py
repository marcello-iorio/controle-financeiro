from flask import Flask, request, jsonify
from db import get_connection
from flask_cors import CORS

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)

def format_moeda(valor):
    return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

@app.route("/")
def home():
    return "API Controle Financeiro funcionando"

@app.route("/movimentacoes", methods=["POST"])
def criar_movimentacao():
    data = request.json

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        INSERT INTO movimentacoes (tipo, valor, data, descricao, categoria_id)
        VALUES (%s, %s, %s, %s, %s)
    """

    values = (
        data["tipo"],
        data["valor"],
        data["data"],
        data.get("descricao"),
        data.get("categoria_id")
    )

    cursor.execute(sql, values)
    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({"status": "ok"}), 201

@app.route("/movimentacoes", methods=["GET"])
def listar_movimentacoes():
    # Captura os filtros da URL
    tipo = request.args.get("tipo")
    categoria_id = request.args.get("categoria_id")
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Base da query
    sql = """
        SELECT m.*, c.nome as categoria 
        FROM movimentacoes m 
        LEFT JOIN categorias c ON m.categoria_id = c.id 
        WHERE 1=1
    """
    params = []

    # FILTRO DE TIPO (Ajuste aqui)
    if tipo and tipo != "todos":
        sql += " AND m.tipo = %s"
        params.append(tipo)

    # Filtro de Categoria
    if categoria_id:
        sql += " AND m.categoria_id = %s"
        params.append(categoria_id)

    # Filtro de Data
    if data_inicio and data_fim:
        sql += " AND m.data BETWEEN %s AND %s"
        params.extend([data_inicio, data_fim])
    elif data_inicio:
        sql += " AND m.data >= %s"
        params.append(data_inicio)
    elif data_fim:
        sql += " AND m.data <= %s"
        params.append(data_fim)

    sql += " ORDER BY m.data DESC"

    cursor.execute(sql, params)
    movimentacoes = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(movimentacoes)

@app.route("/saldo", methods=["GET"])
def calcular_saldo():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            SUM(CASE WHEN tipo = 'entrada' THEN valor ELSE 0 END) -
            SUM(CASE WHEN tipo = 'saida' THEN valor ELSE 0 END)
        AS saldo
        FROM movimentacoes
    """)

    resultado = cursor.fetchone()
    saldo = resultado[0] if resultado[0] is not None else 0

    cursor.close()
    conn.close()

    return jsonify({"saldo": float(saldo)})

@app.route("/categorias", methods=["POST"])
def criar_categoria():
    data = request.json

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        "INSERT INTO categorias (nome) VALUES (%s)",
        (data["nome"],)
    )

    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({"id": cursor.lastrowid,"nome": data["nome"]}), 201

@app.route("/categorias", methods=["GET"])
def listar_categorias():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM categorias")
    categorias = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(categorias)


@app.route("/movimentacoes/<int:id>", methods=["DELETE"])
def excluir_movimentacao(id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM movimentacoes WHERE id = %s", (id,))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"status": "excluido"}), 200

@app.route("/movimentacoes/<int:id>", methods=["PUT"])
def atualizar_movimentacao(id):
    data = request.json

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        UPDATE movimentacoes
        SET tipo = %s,
            valor = %s,
            data = %s,
            descricao = %s,
            categoria_id = %s
        WHERE id = %s
    """

    cursor.execute(sql, (
        data["tipo"],
        data["valor"],
        data["data"],
        data["descricao"],
        data["categoria_id"],
        id
    ))

    conn.commit()

    cursor.close()
    conn.close()

    return jsonify({"status": "atualizado"})



@app.route("/movimentacoes/exportar/excel", methods=["POST"])
def exportar_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from flask import send_file
    from dateutil import parser

    movimentacoes = request.json or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    colunas  = ["Tipo", "Categoria", "Descrição", "Valor (R$)", "Data"]
    larguras = [12, 18, 35, 14, 14]

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    borda = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )

    for col, (nome, larg) in enumerate(zip(colunas, larguras), 1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
        ws.column_dimensions[get_column_letter(col)].width = larg
    ws.row_dimensions[1].height = 22

    fill_par    = PatternFill("solid", fgColor="EFF6FF")
    fill_impar  = PatternFill("solid", fgColor="FFFFFF")
    font_normal  = Font(name="Arial", size=10)
    font_entrada = Font(name="Arial", size=10, color="059669")
    font_saida   = Font(name="Arial", size=10, color="DC2626")

    from datetime import datetime as dt
    for i, m in enumerate(movimentacoes, 2):
        tipo = "Entrada" if m.get("tipo") == "entrada" else "Saída"
        data_raw = m.get("data", "")
        try:
            # data_fmt = dt.strptime(data_raw[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            data_dt = parser.parse(data_raw)
            data_fmt = data_dt.strftime("%d/%m/%Y")
        except:
            data_fmt = data_raw
        linha = [tipo, m.get("categoria") or "Sem categoria", m.get("descricao") or "",
                 float(m.get("valor", 0)), data_fmt]
        fill = fill_par if i % 2 == 0 else fill_impar
        for col, valor in enumerate(linha, 1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.fill   = fill
            cell.border = borda
            cell.alignment = Alignment(vertical="center")
            if col == 1:
                cell.font = font_entrada if tipo == "Entrada" else font_saida
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 4:
                cell.font = font_entrada if tipo == "Entrada" else font_saida
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.font = font_normal
        ws.row_dimensions[i].height = 20

    ultima_linha = len(movimentacoes) + 1
    tabela = Table(displayName="Movimentacoes", ref=f"A1:E{ultima_linha}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tabela)

    entradas = sum(float(m.get("valor", 0)) for m in movimentacoes if m.get("tipo") == "entrada")
    saidas   = sum(float(m.get("valor", 0)) for m in movimentacoes if m.get("tipo") == "saida")
    saldo    = entradas - saidas

    linha_total = ultima_linha + 2
    for offset, (label, valor, cor) in enumerate([
        ("Total de entradas", entradas, "059669"),
        ("Total de saídas",   saidas,   "DC2626"),
        ("Saldo",             saldo,    "1D4ED8"),
    ]):
        ws.cell(row=linha_total + offset, column=1, value=label).font = Font(bold=True, name="Arial", size=10)
        cell_val = ws.cell(row=linha_total + offset, column=4, value=valor)
        cell_val.font = Font(bold=True, name="Arial", size=10, color=cor)
        cell_val.number_format = "#,##0.00"
        cell_val.alignment = Alignment(horizontal="right")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="movimentacoes.xlsx")

@app.route("/movimentacoes/exportar/pdf", methods=["POST"])
def exportar_pdf():
    import io
    from datetime import datetime
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from dateutil import parser

    movimentacoes = request.json or []
    filtros_info  = request.args.get("filtros", "")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("<b>Controle Financeiro — Histórico</b>",
        ParagraphStyle("titulo", fontSize=14, spaceAfter=4)))

    if filtros_info:
        elementos.append(Paragraph(f'<font size=8 color=grey>Filtros: {filtros_info}</font>',
            styles["Normal"]))

    entradas = sum(float(m.get("valor", 0)) for m in movimentacoes if m.get("tipo") == "entrada")
    saidas   = sum(float(m.get("valor", 0)) for m in movimentacoes if m.get("tipo") == "saida")
    saldo    = entradas - saidas
    entradas = format_moeda(entradas)
    saidas = format_moeda(saidas)
    saldo = format_moeda(saldo)

    
    elementos.append(Paragraph(
        f'<font size=9>Saldo: <b>R$ {saldo}</b></font>', styles["Normal"]))
    elementos.append(Spacer(1, 0.4*cm))

    cabecalho = ["Tipo", "Categoria", "Descrição", "Valor (R$)", "Data"]
    linhas = [cabecalho]

    for m in movimentacoes:
        tipo = "Entrada" if m.get("tipo") == "entrada" else "Saída"
        data_raw = m.get("data", "")
        try:
            # data_fmt = datetime.strptime(data_raw[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            # O parser.parse identifica o formato automaticamente (ideal para formatos GMT/UTC)
            data_dt = parser.parse(data_raw)
            data_fmt = data_dt.strftime("%d/%m/%Y")
        except:
            data_fmt = data_raw
        linhas.append([
            tipo,
            m.get("categoria") or "Sem categoria",
            m.get("descricao") or "",
            f'R$ {format_moeda(float(m.get("valor", 0)))}',
            data_fmt
        ])

    col_widths = [2.5*cm, 3.5*cm, 7*cm, 3*cm, 3*cm]
    tabela = Table(linhas, colWidths=col_widths, repeatRows=1)

    estilo = TableStyle([
        ("BACKGROUND",     (0,0), (-1,0),  colors.HexColor("#2563EB")),
        ("TEXTCOLOR",      (0,0), (-1,0),  colors.white),
        ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0,0), (-1,0),  9),
        ("ALIGN",          (0,0), (-1,0),  "CENTER"),
        ("FONTNAME",       (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",       (0,1), (-1,-1), 8),
        ("ALIGN",          (3,1), (3,-1),  "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#EFF6FF"), colors.white]),
        ("GRID",           (0,0), (-1,-1), 0.4, colors.HexColor("#D1D5DB")),
        ("TOPPADDING",     (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 4),
    ])

    for i, m in enumerate(movimentacoes, 1):
        cor = colors.HexColor("#059669") if m.get("tipo") == "entrada" else colors.HexColor("#DC2626")
        estilo.add("TEXTCOLOR", (0,i), (0,i), cor)
        estilo.add("TEXTCOLOR", (3,i), (3,i), cor)

    tabela.setStyle(estilo)
    elementos.append(tabela)
    elementos.append(Spacer(1, 0.6*cm))

    totais_data = [
        ["Total de entradas", "", "", f'R$ {entradas}', ""],
        ["Total de saídas",   "", "", f'R$ {saidas}',  ""],
        ["Saldo",             "", "", f'R$ {saldo}',   ""],
    ]
    tabela_totais = Table(totais_data, colWidths=col_widths)
    estilo_totais = TableStyle([
        ("FONTNAME",        (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",        (0,0), (-1,-1), 9),
        ("ALIGN",           (3,0), (3,-1),  "RIGHT"),
        ("TOPPADDING",      (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",   (0,0), (-1,-1), 3),
    ])
    for i, cor_hex in enumerate(["#059669", "#DC2626", "#1D4ED8"]):
        c = colors.HexColor(cor_hex)
        estilo_totais.add("TEXTCOLOR", (0,i), (0,i), c)
        estilo_totais.add("TEXTCOLOR", (3,i), (3,i), c)
    tabela_totais.setStyle(estilo_totais)
    elementos.append(tabela_totais)

    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, mimetype="application/pdf",
        as_attachment=True, download_name="movimentacoes.pdf")

if __name__ == "__main__":
    app.run(debug=True)